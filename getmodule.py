import requests
from dotenv import load_dotenv
import os
import re
import json
from sentence_transformers import SentenceTransformer, util
import numpy as np
from langchain.prompts import ChatPromptTemplate
from langchain.schema.runnable import RunnableMap
from langchain_community.chat_models import ChatOllama
from langchain.schema.runnable import RunnableMap
import pprint
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain.embeddings.sentence_transformer import SentenceTransformerEmbeddings
from langchain.vectorstores import FAISS
from langchain.docstore.document import Document
import warnings
import itertools
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
# LangChainDeprecationWarning 경고 숨기기
warnings.filterwarnings("ignore", category=UserWarning, module="langchain")

# FutureWarning 경고 숨기기
warnings.filterwarnings("ignore", category=FutureWarning, module="transformers")
warnings.filterwarnings("ignore")

# 다른 경고도 필요에 따라 추가할 수 있습니다.


##gpt 생성시 딕셔너리 외의 단어가 나올 수 있기 때문에 {.*}로 딕셔너리만 파싱한다.

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_market_analysis_report(data, querry):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # 제목 설정 및 중간 굵기 테두리

    sheet['B2'].border = Border(top=Side(style='medium'))
    sheet['C2'].border = Border(top=Side(style='medium'))
    sheet['D2'].border = Border(top=Side(style='medium'))
    sheet['E2'].border = Border(top=Side(style='medium'))
    sheet['F2'].border = Border(top=Side(style='medium'))
    sheet['G2'].border = Border(top=Side(style='medium'))

    sheet['I3'] = "범례"
    sheet['I4'] = "최대값"
    sheet['I5'] = "최소값"

    fill_green = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    sheet['I4'].fill =fill_yellow
    sheet['I5'].fill =fill_green
    alignment_center = Alignment(horizontal='center', vertical='center')
    sheet['I3'].alignment = alignment_center
    sheet['I4'].alignment = alignment_center
    sheet['I5'].alignment = alignment_center

    # I열의 열 너비를 텍스트에 맞춰 자동 조정 (공백 두 칸 추가)
    max_length = 0
    column = 'I'
    for cell in sheet[column]:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2) + 5000  # 추가 공간(공백 두 칸)을 고려하여 열 너비 조정
    sheet.column_dimensions[column].width = adjusted_width



    for idx in range(2,len(data)+2+6,1):
        sheet[f'H{str(idx)}'].border = Border(left=Side(style='medium'))

    for idx in range(2,len(data)+2+6,1):
        sheet[f'A{str(idx)}'].border = Border(right=Side(style='medium'))

    sheet[f'I3'].border = Border(left=Side(style='medium'),right=Side(style='medium'),top=Side(style='medium'))
    sheet[f'I4'].border = Border(left=Side(style='medium'),right=Side(style='medium'))
    sheet[f'I5'].border = Border(left=Side(style='medium'),right=Side(style='medium'),bottom=Side(style='medium'))

    last = f'{len(data)+2+6}'
    last_1 = f'{len(data)+2+5}'

    sheet[f'B6'].border = Border(top=Side(style='medium'))
    sheet[f'C6'].border = Border(top=Side(style='medium'))
    sheet[f'D6'].border = Border(top=Side(style='medium'))
    sheet[f'E6'].border = Border(top=Side(style='medium'))
    sheet[f'F6'].border = Border(top=Side(style='medium'))
    sheet[f'G6'].border = Border(top=Side(style='medium'))

    sheet[f'B6'].border = Border(top=Side(style='medium'))
    sheet[f'C6'].border = Border(top=Side(style='medium'))
    sheet[f'D6'].border = Border(top=Side(style='medium'))
    sheet[f'E6'].border = Border(top=Side(style='medium'))
    sheet[f'F6'].border = Border(top=Side(style='medium'))
    sheet[f'G6'].border = Border(top=Side(style='medium'))

    sheet[f'B{last_1}'].border = Border(top=Side(style='medium'))
    sheet[f'C{last_1}'].border = Border(top=Side(style='medium'))
    sheet[f'D{last_1}'].border = Border(top=Side(style='medium'))
    sheet[f'E{last_1}'].border = Border(top=Side(style='medium'))
    sheet[f'F{last_1}'].border = Border(top=Side(style='medium'))
    sheet[f'G{last_1}'].border = Border(top=Side(style='medium'))

    sheet[f'B{last}'].border = Border(top=Side(style='medium'))
    sheet[f'C{last}'].border = Border(top=Side(style='medium'))
    sheet[f'D{last}'].border = Border(top=Side(style='medium'))
    sheet[f'E{last}'].border = Border(top=Side(style='medium'))
    sheet[f'F{last}'].border = Border(top=Side(style='medium'))
    sheet[f'G{last}'].border = Border(top=Side(style='medium'))

    sheet.merge_cells('B2:G5')
    sheet['B2'] = f'{querry} 제품 시장조사'
    sheet['B2'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['B2'].font = Font(size=16, bold=True)
    ###스타일 시트 생성




    # 셀의 시작 위치 
    start_row = 6
    start_col = 2  # 'B'에 해당

    # 헤더 작성
    headers = ['상품 이름', '브랜드', '제조사', '상품 최고가', '상품 최저가', '상품 링크']
    for col_num, header in enumerate(headers, start=start_col):
        sheet.cell(row=start_row, column=col_num, value=header)

    # 데이터 작성
    for i, row in enumerate(data, start=start_row + 1):
        sheet.cell(row=i, column=start_col, value=row['상품 이름'])
        sheet.cell(row=i, column=start_col + 1, value=row['브랜드'])
        sheet.cell(row=i, column=start_col + 2, value=row['제조사'])
        sheet.cell(row=i, column=start_col + 3, value=int(row['상품 최고가']))
        sheet.cell(row=i, column=start_col + 4, value=int(row['상품 최고가']))
        
        # 상품판매처를 표시하고, 상품 링크를 하이퍼링크로 설정
        if row['상품 링크']:
            sheet.cell(row=i, column=start_col + 5).value = f'=HYPERLINK("{row["상품 링크"]}", "{row["쇼핑몰 정보"]}")'
            sheet.cell(row=i, column=start_col + 5).style = "Hyperlink"
        else:
            sheet.cell(row=i, column=start_col + 5, value="")

    # 평균 행 추가
    average_row = start_row + len(data) + 1
    sheet.cell(row=average_row, column=start_col, value='비고/평균')
    sheet.cell(row=average_row, column=start_col + 3, value=f"=AVERAGE({get_column_letter(start_col+3)}{start_row+1}:{get_column_letter(start_col+3)}{average_row-1})")
    sheet.cell(row=average_row, column=start_col + 4, value=f"=AVERAGE({get_column_letter(start_col+4)}{start_row+1}:{get_column_letter(start_col+4)}{average_row-1})")

    min_lowest_price = min(data, key=lambda x: x['상품 최저가'])['상품 최저가']
    max_lowest_price = max(data, key=lambda x: x['상품 최저가'])['상품 최저가']

    # 상품 최고가의 최솟값과 최댓값 찾기
    min_highest_price = min(data, key=lambda x: x['상품 최고가'])['상품 최고가']
    max_highest_price = max(data, key=lambda x: x['상품 최고가'])['상품 최고가']



    # 셀 스타일링
    currency_format = '"₩ "#,##0'
    fill_green = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    # 각 행에서 최고가와 최저가를 찾아 색상 적용 및 통화 형식 적용
    currency_format = '"₩"#,##0'

    # E 열의 모든 셀에 대해 통화 형식을 적용
    for row in range(7, len(data)+8):
        sheet[f'E{row}'].number_format = currency_format
        #E열이 상품 최고가
    for row in range(7, len(data)+8):
        #F열이 상품 최저가
        sheet[f'F{row}'].number_format = currency_format

    fill_green = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # 초록색 (최솟값)
    fill_yellow = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # 노란색 (최댓값)

    for idx, value in enumerate(data, start=7):
        cell = sheet[f'E{idx}']
        if cell.value == max_highest_price:
            cell.fill = fill_yellow
        elif cell.value == min_highest_price:
            cell.fill = fill_green


    for idx, value in enumerate(data, start=7):
        cell = sheet[f'F{idx}']
        if cell.value == max_lowest_price:
            cell.fill = fill_yellow
        elif cell.value == min_lowest_price:
            cell.fill = fill_green

    fill_red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    for row in range(6, len(data) + 7):
        cell = sheet[f'C{row}']
        cell.alignment = alignment_center
        if cell.value == "":
            cell.fill = fill_red

    for row in range(6, len(data) + 7):
        cell = sheet[f'D{row}']
        cell.alignment = alignment_center
        if cell.value == "":
            cell.fill = fill_red

    for row in range(6, len(data) + 7):
        cell = sheet[f'E{row}']
        if cell.value == "":
            cell.fill = fill_red

    for row in range(6, len(data) + 7):
        cell = sheet[f'G{row}']
        cell.alignment = alignment_center
        if cell.value == "":
            cell.fill = fill_red

    for row in range(6, len(data) + 7):
        cell = sheet[f'B{row}']
        cell.alignment = alignment_center
        if cell.value == "":
            cell.fill = fill_red


    sheet[f'F6'].alignment = alignment_center
    sheet[f'E6'].alignment = alignment_center

    # 테두리 스타일 적용 (데이터 테두리만)
    # 열 너비 자동 조정
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # 열의 첫 번째 셀의 열 문자
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)+3
        sheet.column_dimensions[column].width = adjusted_width

    # 엑셀 파일 저장
    workbook.save("선풍기_제품_시장조사_최종_수정.xlsx")


    # 파일 저장
    filename = f"{querry}_상품_분석_제품_시장조사.xlsx"
    try:
        workbook.save(filename)
        return {f"msg": f"{filename}파일로 성공적으로 저장되었습니다. 사용자 에게는 다음 데이터 리스트를 5열 까지만 정리해서 데이터 프레임으로 보여주세요. 데이터 리스트f{data}"}
    except Exception as e:
        return {"msg": "저장에 실패했습니다."}

load_dotenv()

g_context = {}
total_item = None
def set_context(value):
    global g_context  # g_context를 전역 변수로 사용
    g_context = value

def get_context():
    global g_context 
    return g_context

def set_total_item(value):
    global total_item  # g_context를 전역 변수로 사용
    total_item = value

def get_total_item():
    global total_item 
    return total_item

def set_querr_item(value):
    global querr_item  # g_context를 전역 변수로 사용
    querr_item = value

def get_querr_item():
    global querr_item 
    return querr_item



#쇼핑 디코딩
def parse_shoping_data(data):
    items = data["items"]
    porducts = []
    productType_decoder = {
    "1": "일반상품 가격비교 상품",
    "2": "일반상품 가격비교 비매칭 일반상품",
    "3": "일반상품 가격비교 매칭 일반상품",
    "4": "중고상품 가격비교 상품",
    "5": "중고상품 가격비교 비매칭 일반상품",
    "6": "중고상품 가격비교 매칭 일반상품",
    "7": "단종상품 가격비교 상품",
    "8": "단종상품 가격비교 비매칭 일반상품",
    "9": "단종상품 가격비교 매칭 일반상품",
    "10": "판매예정상품 가격비교 상품",
    "11": "판매예정상품 가격비교 비매칭 일반상품",
    "12": "판매예정상품 가격비교 매칭 일반상품"
}

    for item in items:
        info = {
            '상품 이름': re.sub(r'<[^>]+>', '', item.get('title')),
            '상품 링크': item.get('link'),
            '상품 이미지': item.get('image'),
            '상품 최고가': item.get('hprice') if item.get('hprice') != "" else item.get('lprice') if item.get('lprice') != "" else "0",
            '상품 최저가': item.get('lprice') if item.get('lprice') != "" else item.get('hprice') if item.get('hprice') != "" else "0",
            '쇼핑몰 정보': item.get('mallName'),
            '상품 타입': productType_decoder.get(item.get('productType')),
            '제조사': item.get('maker'),
            '브랜드': item.get('brand'),
            '대분류 카테고리': item.get('category1'),
            '중분류 카테고리': item.get('category2'),
            '소분류 카테고리': item.get('category3'),
            '세분류 카테고리': item.get('category4'),
        }
        porducts.append(info)
    return porducts




def sort_function(query):
    sort_dict =None
    items = get_context()["저장된 아이템 목록"]["items"]
    if query["정렬 쿼리"] == "최고가":
        sort_dict = sorted(items, key=lambda x: int(x["상품 최저가"]))
    elif query["정렬 쿼리"] == "최저가":
        sort_dict = sorted(items, key=lambda x: int(x["상품 최저가"]), reverse=True)
    elif query["정렬 쿼리"] == "검색순":
        sort_dict = items
    len_item = int(query["저장할 개수"])
    return create_market_analysis_report(sort_dict[:len_item],get_querr_item())


    


def make_query(inputs):
    template = """
    너는 사용자의 질문에서 검색 쿼리를 찾는 모델이야.
    사용자의 질문에서 검색 쿼리 키워드를 찾아줘. 
    반드시 JSON으로 반환해줘
    
    반드시 다음 형식으로 대답해줘:
    {{
        "검색 쿼리": "문장에서 뽑아낸 검색 키워드"
    }}
    예시는 다음과 같아:
    사용자: 쇼핑몰에서 최저가 순으로 참치를 10개 검색해줘
    답변: {{"검색 쿼리": "참치"}}

    사용자: 가격이 낮은 컴퓨터 검색해줘
    답변: {{ "검색 쿼리": "컴퓨터"}}

    사용자: 네네치킨을 가격이 낮은 순으로 검색해줘
    답변: {{"검색 쿼리": "네네치킨"}}

    사용자: 인형은 10개를 가격이 낮은 순으로 검색해줘
    답변: {{"검색 쿼리": "인형"}}

    이제 너가 대답해줄 차례야
    사용자: {question} 
    """

    prompt = ChatPromptTemplate.from_template(template)
    llm = ChatOllama(model="gemma2:9b", temperature=0, base_url="http://127.0.0.1:11434/") #http://127.0.0.1:11434
    chain = RunnableMap({
    "question": lambda x: x["question"]
    }) | prompt | llm  
    chat_msg = chain.invoke({'question': f"{inputs}"}).content
    query_dict_re = re.search(r'\{.*?\}', chat_msg, re.DOTALL)
    query_dict = json.loads(query_dict_re.group(0))

    if type(query_dict["검색 쿼리"]) == list:
        return {"msg":"사용자에게 아직은 여러 검색 기능을 지원하지 않아 검색하지 못했다고 답해줘."}
    return {"msg":f"다음 내용을 사용자에게 잘 전달해줘 사용자의 질문에서 {query_dict['검색 쿼리']} 검색 키워드를 찾았습니다.","검색 쿼리":query_dict["검색 쿼리"]}
    #프롬포트에 sort_list를 전달, 쿼리가 나오도록 전달해줌.


def make_url(query:dict):
    serach_query = query["검색 쿼리"]
    pageable = 1
    try:
            headers = {
            "X-Naver-Client-Id": os.getenv("X-Naver-Client-Id"),
            "X-Naver-Client-Secret": os.getenv("X-Naver-Client-Secret")
            }
            response = requests.get(f"https://openapi.naver.com/v1/search/shop.json?query={serach_query}&sort=sim&start_page=1",headers=headers)
            if response.status_code == 200:
                json_f = response.json()
                total = json_f["total"]
            else:
                {"msg":"검색 엔진으로 검색하지 못했습니다."}
            pageable = total // 100
            pageable += 1
            
            if pageable >= 5:
                pageable = 5

            page_list= []
            for idx in range(pageable):
                start_index = idx * 100  
                # display_len 계산
                display_len = total - start_index if total - start_index >= 100 else total - start_index
                
                # display_len이 0이 아닌 경우에만 URL 추가
                if display_len > 0:
                    page_list.append(f"https://openapi.naver.com/v1/search/shop.json?query={serach_query}&sort=sim&start_page={idx + 1}&display_len={display_len}")
            
            
    except Exception as e:
        return {"msg":"검색 쿼리를 생성하지 못했습니다."}
    set_querr_item(serach_query)
    return {"msg":f"다음 너가 할 행동이야:{serach_query}를 통해 {total}개의 검색 결과를 찾았습니다. 이중 {pageable * 100}개의 검색을 시작하겠습니다. 너의 행동에 대해 잘 설명해줘","urls":page_list}

def request_for_serach_engen(querry:dict):
    headers = {
        "X-Naver-Client-Id": os.getenv("X-Naver-Client-Id"),
        "X-Naver-Client-Secret": os.getenv("X-Naver-Client-Secret")
    }
    baseurl = querry["urls"]
    items = []
    for url in baseurl:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            # 각 URL에 대한 응답을 파싱해서 items에 추가
            items.append(parse_shoping_data(response.json()))
        elif response.status_code == 401:
            return {"msg": "서버 에러, 관리자에게 문의하세요(API 키 오류)"}
        elif response.status_code == 429:
            return {"msg": "서버 에러, 관리자에게 문의하세요(API 키 요청한도 초과)"}
        else:
            continue
    items = list(itertools.chain(*items))
    return {"msg":"검색 성공! 검색 결과를 저장하겠습니다.","items":items,"total":response.json()["total"]}

def pick_sentence(inputs):
    if "저장된 아이템 목록" not in  get_context():
        context_is = False
    else:
        context_is = len(get_context()) == 0

    text = "없음" if context_is else "문서작성"
    text2 = "탁구공에 대한 검색을 먼저 실행해 주세요" if context_is else "네 주어진 검색 결과를 통해 탁구공에 대해 가격이 낮은 순으로 정렬해서 엑셀로 정리해 드리겠습니다."
    human = "없음" if context_is else "정리"
    human_sentence = "빽다방 쿠폰에 대해 검색을 먼저 실행 해 주세요." if context_is else "네 주어진 검색 결과에서 빽다방 쿠폰이 얼마인지 검색해 드리겠습니다."
    orderlist=["검색","문서작성","정리"]
    template = """
    너는 사용자의 질문에서 실행할 명령을 찾는 모델이야. 
    명령 리스트를 충분하게 확인하고 사용자의 질문중 해당하는 명령이 있으면 반드시 사람이 읽을 수 있는 형태와 JSON 형태로 반환해줘.
    검색은 반드시 쇼핑몰에서만 검색하기 때문에 쇼핑몰에서 검색 하겠습니다 라는 말을 포함해서 반환해줘야해.
    JSON을 이용해 명령을 실행시킬것이기 때문에 반드시 JSON도 반환해줘야해.
    만약 사용자의 질문에서 명령 리스트를 찾지 못하면, 실행 할 수 없는 명령입니다 라는 말로 반환해줘.
    엑셀이라는 키워드가 검색이 아닌 명령 키워드중 포함된다면 무조건 문서작성 키워드로 반환해줘.
    명령 리스트는 다음과 같아.
    {orderlist}
    {context_is} 이 값이 True면 문서작성 과 정리 명령은 무조건 없음 명령으로 출력하고 문서로 작성할 데이터를 먼저 검색해 주세요 라는 메세지를 출력해줘.
    
    JSON 형식은 다음과 같아.
    {{
        "명령":"명령 리스트 중 한개의 값"
    }}

    대답의 예시를 들어줄게.

    사용자의 질문: 유튜브 뮤직에서 음악 틀어줘
    답장: 저는 유튜브 뮤직에서 음악을 재생할 수 없는 모델입니다. 다른 도움을 드릴까요?
    {{
        "명령":"없음"
    }}
    
    사용자의 질문: 쇼핑몰에서 아기 옷 검색해줘.
    답장: 네 쇼핑몰에서 아기옷을 검색해 드리겠습니다.
    {{
        "명령":"검색"
    }}

    사용자의 질문: 검색된 내용에서 빽다방 쿠폰은 얼마야?
    답장: {human_sentence}
    {{
        "명령":"{human}"
    }} 

    사용자의 질문: 검색된 내용에서 탁구공 의 가격을 엑셀로 정리해줘
    답장: {text2}
    {{
        "명령":"{text}"
    }} 

    사용자의 질문: 검색된 내용에서 탁구공 가격을 가격이 낮은순으로 정렬해서 엑셀로 저장해줘
    답장: {text2}
    {{
        "명령":"{text}"
    }} 
    
    이제 너의 차례야, 사용자의 질문에 대답해줘
    사용자의 질문 {inputs}
    """
    prompt = ChatPromptTemplate.from_template(template)
    llm = ChatOllama(model="gemma2:9b", temperature=0, base_url="http://127.0.0.1:11434/") #http://127.0.0.1:11434
    chain = RunnableMap({
    "orderlist": lambda x: str(x["orderlist"]),
    "inputs": lambda x: x["inputs"],
    "context_is": lambda x: x["context_is"],
    "human_sentence": lambda x: x["human_sentence"],
    "human": lambda x: x["human"],
    "text2": lambda x: x["text2"],
    "text": lambda x: x["text"],
    }) | prompt | llm 
    chat_msg = chain.invoke({"orderlist":orderlist,"inputs":inputs,"context_is":context_is,"human_sentence":human_sentence,"human":human,"text2":text2,"text":text}).content
    order_dict_re = re.search(r'\{.*?\}', chat_msg, re.DOTALL)
    human_read = chat_msg.replace(order_dict_re.group(0),"")
    human_read = re.sub(r'^\s*\n', '', human_read, flags=re.MULTILINE)
    print(human_read) #TODO 이거 스트림릿 채팅에 나오게 해주세요
    order_dict = json.loads(order_dict_re.group(0))
    return order_dict

def make_sentence(inputs):
    text_splitter = RecursiveCharacterTextSplitter(
    chunk_size=1000, 
    chunk_overlap=200
)
    texts = text_splitter.split_text(json.dumps(get_context(), ensure_ascii=False, indent=2))
    documents = [Document(page_content=text) for text in texts]
    embedding_function = SentenceTransformerEmbeddings(model_name="jhgan/ko-sroberta-multitask")
    db = FAISS.from_documents(documents, embedding_function)
    template = """
        너는 결과 정리 봇이야. 먼저 차근차근 다음 맥락을 읽어줘.
        
        맥락

        {context}

        반드시 한글로 대답해줘.
        이후 사용자의 요구에 맞춰 맥락에 있는 내용으로 사용자에게 필요한 정보를 제공해줘.
        
        사용자의 요구는 다음과 같아.
        {question}
        
        이제 대답을 해줘 반드시 한글로 작성해야해
        """
    
    llm = ChatOllama(model="gemma2:9b", temperature=0, base_url="http://127.0.0.1:11434/") #http://127.0.0.1:11434
    prompt = ChatPromptTemplate.from_template(template)
    retriever = db.as_retriever(search_type="similarity", search_kwargs={'k':10, 'fetch_k': 100})
    chain = RunnableMap({
    "context": lambda x: retriever.get_relevant_documents(x['question']),
    "question": lambda x: x['question']
    }) | prompt | llm

    return chain.invoke({'question': f"{inputs}"}).content

def pick_sort_query(inputs):
    sort_list = [
        "최고가",
        "최저가",
        "검색순"
    ]
    template = """
        너는 사용자의 질문에서 정렬 쿼리와 저장할 갯수를 찾는 모델이야.
        반드시 JSON 형식으로 대답해줘.
        정렬 쿼리 리스트는 다음과 같아
        {sort_list}
        정렬 쿼리는 반드시 정렬 쿼리 리스트 내에서 반환해줘.
        만약 사용자의 질문에서 정렬 쿼리를 찾지 못한다면 키 값만 형식에 맞게 채워주고 값에는 ""로 채워줘
        만약 사용자의 질문에서 저장할 개수를 찾지 못한다면 키 값만 형식에 맞게 채워주고 값에는 ""로 채워줘

        반드시 다음 형식으로 대답해줘:
        {{
            "정렬 쿼리": "정렬쿼리 리스트 내의 값",
            "저장할 개수": "문장에서 뽑아낸 저장할 개수"
        }}
        예시는 다음과 같아:
        사용자: 검색된 아이템을 최저가 순으로 10개 저장해줘
        답변: {{"정렬 쿼리": "최저가", "저장할 개수": 10}}

        사용자: 네네치킨 검색결과를 최고가 순으로 20개 저장해줘
        답변: {{"정렬 쿼리": "최고가", "저장할 개수": 20}}
        
        사용자: 검색결과를 검색순으로 5개 저장해줘
        답변: {{"정렬 쿼리": "검색순", "저장할 개수": 5}}

        이제 너가 대답해줄 차례야
        사용자: {question} 
        """
    
    prompt = ChatPromptTemplate.from_template(template)
    llm = ChatOllama(model="gemma2:9b", temperature=0, base_url="http://127.0.0.1:11434/") #http://127.0.0.1:11434
    chain = RunnableMap({
    "sort_list": lambda x: x["sort_list"],
    "question": lambda x: x["question"]
    }) | prompt | llm  
    chat_msg = chain.invoke({'question': f"{inputs}","sort_list":f"{',' .join(sort_list)}"}).content
    sort_re = re.search(r'\{.*?\}', chat_msg, re.DOTALL)
    order_dict = json.loads(sort_re.group(0))
    msg = ""
    err_msg=""
    err_msg_save_len=""
    if order_dict["정렬 쿼리"] == "":
        order_dict["정렬 쿼리"] = "검색순"
        err_msg = "정렬 기준이 없어 검색순으로 정렬하겠습니다."
    if order_dict["저장할 개수"] == "":
        order_dict["저장할 개수"] = 10
        err_msg_save_len = "저장 개수가 없어 10개를 저장하겠습니다."
    msg += "" if order_dict['정렬 쿼리'] == "" else f"{order_dict['정렬 쿼리']}순으로"
    msg += "" if order_dict["저장할 개수"] == "" else f"{order_dict['저장할 개수']}만큼 정리해 Excel 파일로 저장하겠습니다."+err_msg+err_msg_save_len
    return {"msg":msg,"정렬 쿼리":order_dict["정렬 쿼리"],"저장할 개수":order_dict["저장할 개수"]}

def select_sentence(inputs, order:dict):
    template = """
    다음은 너가 사용자에게 전달할 말이야.
    {context}
    잘 전달해주고 부연 설명은 하지 말아줘
    """
    llm = ChatOllama(model="gemma2:9b", temperature=0, base_url="http://127.0.0.1:11434/") #http://127.0.0.1:11434
    prompt = ChatPromptTemplate.from_template(template)
    chain = RunnableMap({
        "context": lambda x: (x['context']),
        }) | prompt | llm
    #["검색","문서작성","정리"]
    if order["명령"] == "검색":
        querry =  make_query(inputs)
        
        chat_msg = chain.invoke({'context': f"{querry['msg']}"}).content
        print(chat_msg) #TODO:이거 스트림릿 채팅창에 나오게 해주세요
        #여기까지 쿼리 만드는 구간
        if "검색 쿼리" not in querry:
            return 
        urls = make_url(querry)
        chat_msg = chain.invoke({'context': f"{urls['msg']}"}).content
        print(chat_msg) #TODO:이거 스트림릿 채팅창에 나오게 해주세요
        item = request_for_serach_engen(urls)
        chat_msg = chain.invoke({'context': f"{item['msg']}"}).content
        print(chat_msg) #TODO:이거 스트림릿 채팅창에 나오게 해주세요
        if 'items' not in item:
            return
        set_context({"JSON에 저장된 아이템의 총 개수":f"{item['total']}","저장된 아이템 목록":item})
    elif order["명령"] == "문서작성":
        query = pick_sort_query(inputs)
        chat_msg = chain.invoke({'context': f"다음 내용을 사용자에게 잘 전달해줘 사용자의 질문에서 {query['msg']}"}).content#TODO:이거 스트림릿 채팅창에 나오게 해주세요
        print(chat_msg) #TODO:이거 스트림릿 채팅창에 나오게 해주세요
        sort_result = sort_function(query)
        chat_msg = chain.invoke({'context': f"너는 지금 성공적으로 엑셀을 저장했어. 메세지를 사용자 화면에 띄워줘{sort_result['msg']}"}).content#TODO:이거 스트림릿 채팅창에 나오게 해주세요
        print(chat_msg) #TODO:이거 스트림릿 채팅창에 나오게 해주세요
    elif order["명령"] == "정리":
        
        print(make_sentence(inputs))
    else:
        return




inputs = "빽다방 기프티콘을 검색해줘"
a = pick_sentence(inputs)
select_sentence(inputs,a)

# inputs = "검색된 내용중 첫번째 내용에 대해 알려줘"
# print("사용자 질문:",inputs)
# a = pick_sentence(inputs)
# select_sentence(inputs,a)
 
inputs = "검색된 내용을 엑셀로 정리해줘"
a = pick_sentence(inputs)
select_sentence(inputs,a)

# query = make_query("삼성 갤럭시 워치3 40개 가격 내림차순으로 정리해서 엑셀로 저장해줘")
# return_url = make_url(query)
# # print(return_url)
# context = request_for_serach_engen(return_url)

# print(make_sentence(context))
# query = make_query("조용필 앨범 최저가 검색해줘")


# 
# response = requests.get(base_url,headers=headers)
# # print(response.content.decode("utf-8"))

# modified_text = re.sub(r'<.*?>', '', response.content.decode("utf-8"))
# print(modified_text)
# print(parse_shoping_data(json.loads(modified_text)))
