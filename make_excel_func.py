import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_market_analysis_report(data, querry):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # 제목 설정 및 중간 굵기 테두리

    sheet['B2'].border = Border(top=Side(style='medium'))
    sheet['C2'].border = Border(top=Side(style='medium'))
    sheet['D2'].border = Border(top=Side(style='medium'))
    sheet['E2'].border = Border(top=Side(style='medium'))
    sheet['F2'].border = Border(top=Side(style='medium'))
    sheet['G2'].border = Border(top=Side(style='medium'))

    sheet['I3'] = "범례"
    sheet['I4'] = "최대값"
    sheet['I5'] = "최소값"

    fill_green = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    sheet['I4'].fill =fill_yellow
    sheet['I5'].fill =fill_green
    alignment_center = Alignment(horizontal='center', vertical='center')
    sheet['I3'].alignment = alignment_center
    sheet['I4'].alignment = alignment_center
    sheet['I5'].alignment = alignment_center

    # I열의 열 너비를 텍스트에 맞춰 자동 조정 (공백 두 칸 추가)
    max_length = 0
    column = 'I'
    for cell in sheet[column]:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2) + 5000  # 추가 공간(공백 두 칸)을 고려하여 열 너비 조정
    sheet.column_dimensions[column].width = adjusted_width



    for idx in range(2,len(data)+2+6,1):
        sheet[f'H{str(idx)}'].border = Border(left=Side(style='medium'))

    for idx in range(2,len(data)+2+6,1):
        sheet[f'A{str(idx)}'].border = Border(right=Side(style='medium'))

    sheet[f'I3'].border = Border(left=Side(style='medium'),right=Side(style='medium'),top=Side(style='medium'))
    sheet[f'I4'].border = Border(left=Side(style='medium'),right=Side(style='medium'))
    sheet[f'I5'].border = Border(left=Side(style='medium'),right=Side(style='medium'),bottom=Side(style='medium'))

    last = f'{len(data)+2+6}'
    last_1 = f'{len(data)+2+5}'

    sheet[f'B6'].border = Border(top=Side(style='medium'))
    sheet[f'C6'].border = Border(top=Side(style='medium'))
    sheet[f'D6'].border = Border(top=Side(style='medium'))
    sheet[f'E6'].border = Border(top=Side(style='medium'))
    sheet[f'F6'].border = Border(top=Side(style='medium'))
    sheet[f'G6'].border = Border(top=Side(style='medium'))

    sheet[f'B6'].border = Border(top=Side(style='medium'))
    sheet[f'C6'].border = Border(top=Side(style='medium'))
    sheet[f'D6'].border = Border(top=Side(style='medium'))
    sheet[f'E6'].border = Border(top=Side(style='medium'))
    sheet[f'F6'].border = Border(top=Side(style='medium'))
    sheet[f'G6'].border = Border(top=Side(style='medium'))

    sheet[f'B{last_1}'].border = Border(top=Side(style='medium'))
    sheet[f'C{last_1}'].border = Border(top=Side(style='medium'))
    sheet[f'D{last_1}'].border = Border(top=Side(style='medium'))
    sheet[f'E{last_1}'].border = Border(top=Side(style='medium'))
    sheet[f'F{last_1}'].border = Border(top=Side(style='medium'))
    sheet[f'G{last_1}'].border = Border(top=Side(style='medium'))

    sheet[f'B{last}'].border = Border(top=Side(style='medium'))
    sheet[f'C{last}'].border = Border(top=Side(style='medium'))
    sheet[f'D{last}'].border = Border(top=Side(style='medium'))
    sheet[f'E{last}'].border = Border(top=Side(style='medium'))
    sheet[f'F{last}'].border = Border(top=Side(style='medium'))
    sheet[f'G{last}'].border = Border(top=Side(style='medium'))

    sheet.merge_cells('B2:G5')
    sheet['B2'] = f'{querry} 제품 시장조사'
    sheet['B2'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['B2'].font = Font(size=16, bold=True)
    ###스타일 시트 생성




    # 셀의 시작 위치 
    start_row = 6
    start_col = 2  # 'B'에 해당

    # 헤더 작성
    headers = ['상품 이름', '브랜드', '제조사', '상품 최고가', '상품 최저가', '상품링크']
    for col_num, header in enumerate(headers, start=start_col):
        sheet.cell(row=start_row, column=col_num, value=header)

    # 데이터 작성
    for i, row in enumerate(data, start=start_row + 1):
        sheet.cell(row=i, column=start_col, value=row['상품 이름'])
        sheet.cell(row=i, column=start_col + 1, value=row['브랜드'])
        sheet.cell(row=i, column=start_col + 2, value=row['제조사'])
        sheet.cell(row=i, column=start_col + 3, value=row['상품 최고가'])
        sheet.cell(row=i, column=start_col + 4, value=row['상품 최저가'])
        
        # 상품판매처를 표시하고, 상품링크를 하이퍼링크로 설정
        if row['상품링크']:
            sheet.cell(row=i, column=start_col + 5).value = f'=HYPERLINK("{row["상품링크"]}", "{row["상품판매처"]}")'
            sheet.cell(row=i, column=start_col + 5).style = "Hyperlink"
        else:
            sheet.cell(row=i, column=start_col + 5, value="")

    # 평균 행 추가
    average_row = start_row + len(data) + 1
    sheet.cell(row=average_row, column=start_col, value='비고/평균')
    sheet.cell(row=average_row, column=start_col + 3, value=f"=AVERAGE({get_column_letter(start_col+3)}{start_row+1}:{get_column_letter(start_col+3)}{average_row-1})")
    sheet.cell(row=average_row, column=start_col + 4, value=f"=AVERAGE({get_column_letter(start_col+4)}{start_row+1}:{get_column_letter(start_col+4)}{average_row-1})")

    min_lowest_price = min(data, key=lambda x: x['상품 최저가'])['상품 최저가']
    max_lowest_price = max(data, key=lambda x: x['상품 최저가'])['상품 최저가']

    # 상품 최고가의 최솟값과 최댓값 찾기
    min_highest_price = min(data, key=lambda x: x['상품 최고가'])['상품 최고가']
    max_highest_price = max(data, key=lambda x: x['상품 최고가'])['상품 최고가']



    # 셀 스타일링
    currency_format = '"₩ "#,##0'
    fill_green = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    # 각 행에서 최고가와 최저가를 찾아 색상 적용 및 통화 형식 적용
    currency_format = '"₩"#,##0'

    # E 열의 모든 셀에 대해 통화 형식을 적용
    for row in range(7, len(data)+8):
        sheet[f'E{row}'].number_format = currency_format
        #E열이 상품 최고가
    for row in range(7, len(data)+8):
        #F열이 상품 최저가
        sheet[f'F{row}'].number_format = currency_format

    fill_green = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # 초록색 (최솟값)
    fill_yellow = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # 노란색 (최댓값)

    for idx, value in enumerate(data, start=7):
        cell = sheet[f'E{idx}']
        if cell.value == max_highest_price:
            cell.fill = fill_yellow
        elif cell.value == min_highest_price:
            cell.fill = fill_green


    for idx, value in enumerate(data, start=7):
        cell = sheet[f'F{idx}']
        if cell.value == max_lowest_price:
            cell.fill = fill_yellow
        elif cell.value == min_lowest_price:
            cell.fill = fill_green

    fill_red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    for row in range(6, len(data) + 7):
        cell = sheet[f'C{row}']
        cell.alignment = alignment_center
        if cell.value == "":
            cell.fill = fill_red

    for row in range(6, len(data) + 7):
        cell = sheet[f'D{row}']
        cell.alignment = alignment_center
        if cell.value == "":
            cell.fill = fill_red

    for row in range(6, len(data) + 7):
        cell = sheet[f'E{row}']
        if cell.value == "":
            cell.fill = fill_red

    for row in range(6, len(data) + 7):
        cell = sheet[f'G{row}']
        cell.alignment = alignment_center
        if cell.value == "":
            cell.fill = fill_red

    for row in range(6, len(data) + 7):
        cell = sheet[f'B{row}']
        cell.alignment = alignment_center
        if cell.value == "":
            cell.fill = fill_red


    sheet[f'F6'].alignment = alignment_center
    sheet[f'E6'].alignment = alignment_center

    # 테두리 스타일 적용 (데이터 테두리만)
    # 열 너비 자동 조정
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # 열의 첫 번째 셀의 열 문자
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)+3
        sheet.column_dimensions[column].width = adjusted_width

    # 엑셀 파일 저장
    workbook.save("선풍기_제품_시장조사_최종_수정.xlsx")


    # 파일 저장
    filename = f"{querry}_상품_분석_제품_시장조사.xlsx"
    try:
        workbook.save(filename)
        return {"msg": "성공적으로 저장되었습니다.", "filename": filename}
    except Exception as e:
        return {"msg": "저장에 실패했습니다.", "error": str(e)}

# 함수 호출 예제
data = [
    {'상품 이름': '샤오미 선풍기', '브랜드': "샤오미", "제조사": "샤오미", "상품 최고가": 1231231, "상품 최저가": 3123, "상품링크": ""},
    {'상품 이름': '샤오미 선풍기', '브랜드': "샤오미", "제조사": "샤오미", "상품 최고가": 1231231, "상품 최저가": 3123, "상품링크": ""},
    {'상품 이름': '샤오미 선풍기', '브랜드': "샤오미", "제조사": "샤오미", "상품 최고가": 1231231, "상품 최저가": 3123, "상품링크": ""},
    {'상품 이름': '샤오미 선풍기', '브랜드': "샤오미", "제조사": "샤오미", "상품 최고가": 1231231, "상품 최저가": 3123, "상품링크": ""},
    {'상품 이름': '샤오미 선풍기', '브랜드': "샤오미", "제조사": "샤오미", "상품 최고가": 1231231, "상품 최저가": 3123, "상품링크": ""},
    {'상품 이름': '샤오미 에어컨', '브랜드': "", "제조사": "", "상품 최고가": 123123, "상품 최저가": 122332, "상품링크": "www.naver.com","상품판매처":"네이버 마켓"},
    {'상품 이름': '샤오미 냉방', '브랜드': "샤오미", "제조사": "삼성", "상품 최고가": 12312323, "상품 최저가": 2131321, "상품링크": "www.naver.com","상품판매처":"G마켓"},
    {'상품 이름': '삼성 선풍기', '브랜드': "LG", "제조사": "삼성", "상품 최고가": 1231, "상품 최저가": 213132122, "상품링크": "www.naver.com","상품판매처":"G마켓"},
    {'상품 이름': '삼성 선풍기', '브랜드': "LG", "제조사": "삼성", "상품 최고가": 1231, "상품 최저가": 213132122, "상품링크": "www.naver.com","상품판매처":"G마켓"}
]
print(create_market_analysis_report(data,"선풍기"))